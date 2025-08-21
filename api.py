import frappe
from frappe import _
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import MergedCell
import io
import json
import zipfile
import os


EXCEL_TEXT_LIMIT = 32767  # Excel max chars per cell


def _clean_value(v):
    """Normalize values for Excel cells."""
    if v is None:
        return ""
    # keep numbers as numbers
    if isinstance(v, (int, float)):
        return v
    # stringify, strip, and cap to Excel's limit
    s = str(v).strip()
    if len(s) > EXCEL_TEXT_LIMIT:
        s = s[:EXCEL_TEXT_LIMIT]
    return s


@frappe.whitelist()
def get_instrumentation_files_excel():
    """
    API TO DOWNLOAD THE INSTRUMENTATION FILE
    """
    # Print the payload received
    frappe.msgprint(f"Payload: {frappe.local.form_dict}")
    print("Payload:", frappe.local.form_dict)

    payload = frappe.local.form_dict

    # Get the ID from payload
    doc_id = payload.get("name")

    if not doc_id:
        frappe.throw("Document ID is required")

    # Fetch the instrumentation data from the document
    doc = frappe.get_doc("Instrumentation Files", doc_id)

    # Print the document name for debuggings
    print("Document Name:", doc.name)
    frappe.msgprint(f"Document: {doc.name}")

    # Fetch the instrumentation data
    instrumentation_data = []
    if hasattr(doc, "instrumentation_output_data") and doc.instrumentation_output_data:
        try:
            instrumentation_data = json.loads(doc.instrumentation_output_data)
            print(
                "Instrumentation data fetched successfully:",
                len(instrumentation_data),
                "records",
            )
            frappe.msgprint(
                f"Data fetched successfully: {len(instrumentation_data)} records"
            )
        except Exception as e:
            frappe.msgprint(f"Error parsing instrumentation data: {str(e)}")
            instrumentation_data = []
    else:
        frappe.msgprint("No instrumentation data found")
        print("No instrumentation data found")

    # Fetch the valve data
    valve_data = []
    if hasattr(doc, "valve_output_data") and doc.valve_output_data:
        try:
            valve_data = json.loads(doc.valve_output_data)
            print(
                "Valve data fetched successfully:",
                len(valve_data),
                "records",
            )
            frappe.msgprint(
                f"Valve data fetched successfully: {len(valve_data)} records"
            )
        except Exception as e:
            frappe.msgprint(f"Error parsing valve data: {str(e)}")
            valve_data = []
    else:
        frappe.msgprint("No valve data found")
        print("No valve data found")

    # Create a zip file containing both Excel files
    zip_buffer = io.BytesIO()

    instrumentation_excel = generate_instrumentation_excel(instrumentation_data)
    valve_excel = generate_valve_excel(valve_data)

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Generate instrumentation Excel file
        zip_file.writestr(
            "instrumentation_schedule.xlsx", instrumentation_excel.getvalue()
        )
        # Generate valve Excel file
        zip_file.writestr("valve_list.xlsx", valve_excel.getvalue())

        # Add the attached DWG file if it exists
        if doc.instrumentation_dwg_file_path:
            try:
                # Get the file content from the attached file
                file_doc = frappe.get_doc(
                    "File", {"file_url": doc.instrumentation_dwg_file_path}
                )
                file_content = file_doc.get_content()

                # Extract the filename from the path
                filename = os.path.basename(doc.instrumentation_dwg_file_path)

                # Add the file to the zip
                zip_file.writestr(filename, file_content)
            except Exception as e:
                frappe.msgprint(f"Could not add DWG file to zip: {str(e)}")
                print(f"Could not add DWG file to zip: {str(e)}")

    # Return the zip file
    zip_buffer.seek(0)
    frappe.local.response.filename = "instrumentation_files.zip"
    frappe.local.response.filecontent = zip_buffer.getvalue()
    frappe.local.response.type = "binary"

    return _("Files generated successfully.")


def generate_instrumentation_excel(data):
    """Generate the instrumentation Excel file safely on a template."""
    template_path = frappe.get_site_path(
        "private", "files", "Instrument Schedule 1.xlsx"
    )

    wb = load_workbook(template_path)
    ws = wb.active

    # You moved start row to 8 – keep it
    start_row = 7

    # Columns you write to (used for unmerge detection)
    target_cols = {
        "A",
        "B",
        "C",
        "E",
        "K",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "Z",
        "AJ",
        "AY",
    }
    target_col_indexes = {column_index_from_string(c) for c in target_cols}

    # 1) Unmerge any merged ranges that intersect your data region & target columns
    #    This prevents: AttributeError: 'MergedCell' object attribute 'value' is read-only
    for mr in list(ws.merged_cells.ranges):
        touches_rows = mr.max_row >= start_row
        touches_cols = any(mr.min_col <= ci <= mr.max_col for ci in target_col_indexes)
        if touches_rows and touches_cols:
            ws.unmerge_cells(str(mr))

    # 2) Populate data
    for i, item in enumerate(data or []):
        row = start_row + i

        ws[f"A{row}"] = _clean_value(i + 1)
        ws[f"B{row}"] = _clean_value(item.get("TAG_NO", ""))
        ws[f"C{row}"] = _clean_value(item.get("TAG_DESCRIPTION", ""))
        ws[f"E{row}"] = _clean_value(item.get("DISP_TEXT_0", ""))

        # Numeric-ish process values: try to coerce to float if they look like numbers
        def num(key):
            v = item.get(key, "")
            # allow strings like "123.45" to become numbers; else keep as cleaned text
            try:
                if isinstance(v, str):
                    v = v.replace(",", "").strip()
                return float(v)
            except Exception:
                return _clean_value(v)

        ws[f"P{row}"] = num("OPERATING*PRESSURE*KG*CM2")
        ws[f"Q{row}"] = num("DESIGN*PRESSURE*KG*CM2")
        ws[f"R{row}"] = num("OPERATING*TEMPERATURE*C")
        ws[f"S{row}"] = num("DESIGN*TEMPERATURE*C")
        ws[f"T{row}"] = num("OPERATING*FLOW*M3*HR")
        ws[f"U{row}"] = num("DESIGN*FLOW*M3*HR")

        ws[f"K{row}"] = _clean_value(item.get("PIPE*MATERIAL", ""))
        ws[f"AJ{row}"] = _clean_value(item.get("MOUNTING_TYPE", ""))
        ws[f"Z{row}"] = _clean_value(item.get("TANK*HEIGHT*MM", ""))
        ws[f"AY{row}"] = _clean_value(item.get("REMARK", ""))

        # ws[f"D{row}"] = _clean_value(item.get("something", ""))
        ws[f"F{row}"] = _clean_value("NA")
        ws[f"J{row}"] = _clean_value("NA")
        ws[f"V{row}"] = _clean_value("NA")
        ws[f"AA{row}"] = _clean_value("NA")
        ws[f"AB{row}"] = _clean_value("NA")
        ws[f"AC{row}"] = _clean_value("NA")
        ws[f"AD{row}"] = _clean_value("NA")
        ws[f"AE{row}"] = _clean_value("NA")
        ws[f"AF{row}"] = _clean_value("NA")
        ws[f"AG{row}"] = _clean_value("NA")
        ws[f"AH{row}"] = _clean_value("NA")
        ws[f"AI{row}"] = _clean_value("NA")
        ws[f"AO{row}"] = _clean_value("NA")
        ws[f"AP{row}"] = _clean_value("NA")
        ws[f"AQ{row}"] = _clean_value("NA")
        ws[f"AR{row}"] = _clean_value("NA")
        ws[f"AS{row}"] = _clean_value("NA")
        ws[f"AV{row}"] = _clean_value("NA")
        ws[f"AW{row}"] = _clean_value("NA")
        ws[f"AX{row}"] = _clean_value("NA")

    # 3) Save to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def generate_valve_excel(data):
    """Generate the valve Excel file using the same structure as instrumentation data"""
    template_path = frappe.get_site_path("private", "files", "Valve List 1 1.xlsx")
    # Load the template workbook
    template_workbook = load_workbook(template_path)

    # Get the first sheet (or the appropriate sheet for valve data)
    worksheet = template_workbook.active

    # Define the starting row for data
    start_row = 7

    # Columns to unmerge that might conflict with our data
    target_cols = {"A", "B", "C", "D", "E", "F"}
    target_col_indexes = {column_index_from_string(c) for c in target_cols}

    # Unmerge any merged ranges that intersect our data region
    for mr in list(worksheet.merged_cells.ranges):
        touches_rows = mr.max_row >= start_row
        touches_cols = any(mr.min_col <= ci <= mr.max_col for ci in target_col_indexes)
        if touches_rows and touches_cols:
            worksheet.unmerge_cells(str(mr))

    # Populate the data into the Excel sheet
    for i, item in enumerate(data):
        row = start_row + i
        # Populate each field in the corresponding columns based on the test data structure

        worksheet[f"A{row}"] = _clean_value(i + 1)
        worksheet[f"B{row}"] = _clean_value(item.get("VALVE_TAG", ""))
        worksheet[f"C{row}"] = _clean_value(item.get("VALVE_TYPE", ""))

        worksheet[f"D{row}"] = _clean_value("NA")
        worksheet[f"E{row}"] = _clean_value("NA")
        worksheet[f"F{row}"] = _clean_value("NA")

        worksheet[f"J{row}"] = _clean_value(item.get("OPERATING PRESSURE (KG/CM2)", ""))
        worksheet[f"K{row}"] = _clean_value(item.get("DESIGN PRESSURE (KG/CM2)", ""))
        worksheet[f"P{row}"] = _clean_value(item.get("OPERATING TEMPERATURE (°C)", ""))
        worksheet[f"Q{row}"] = _clean_value(item.get("DESIGN TEMPERATURE (°C)", ""))
        worksheet[f"R{row}"] = _clean_value(item.get("OPERATING FLOW (M3/HR)", ""))
        worksheet[f"S{row}"] = _clean_value(item.get("DESIGN FLOW (M3/HR)", ""))
        worksheet[f"T{row}"] = _clean_value(item.get("PRESSURE CLASS", ""))
        worksheet[f"U{row}"] = _clean_value(item.get("END CONNECTION", ""))
        worksheet[f"V{row}"] = _clean_value(item.get("BODY MOC", ""))
        worksheet[f"Z{row}"] = _clean_value(item.get("DESIGN STANDARD", ""))
        worksheet[f"AA{row}"] = _clean_value(item.get("OPERATOR / ACTUATION", ""))
    # Add more fields as needed based on the actual template structure

    # Save the workbook to a BytesIO object
    output = io.BytesIO()
    template_workbook.save(output)
    output.seek(0)

    return output
